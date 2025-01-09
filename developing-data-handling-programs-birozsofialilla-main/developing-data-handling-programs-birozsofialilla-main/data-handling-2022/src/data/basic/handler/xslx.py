from typing import Type, cast

import openpyxl
from openpyxl.workbook import Workbook

from data.basic.generator import generate_people
from data.basic.model_dataclasses import Person, Motorbike, Street
from model_dataclasses import Motorbike, Street


def write_people(people: list[Person], wb: openpyxl.Workbook,
                 sheet_name: str = "people",
                 heading: bool = True) -> None:
    sheet = wb.create_sheet(sheet_name)

    if heading:
        column_names = ["id", "name", "age", "male"]
        for col in range(len(column_names)):
            sheet.cell(row=1, column=col + 1, value=column_names[col])

    offset = 2 if heading else 1
    for row in range(len(people)):
        sheet.cell(row=row + offset, column=1, value=people[row].id)
        sheet.cell(row=row + offset, column=2, value=people[row].name)
        sheet.cell(row=row + offset, column=3, value=people[row].age)
        sheet.cell(row=row + offset, column=4, value=people[row].male)



def read_people(wb: Workbook,
                sheet_name: str = "people",
                heading: bool = True) -> list[Person]:
    sheet = wb[sheet_name]

    people = []
    row = 2 if heading else 1
    while True:
        cell = sheet.cell(row=row, column=1)
        if cell.value is None:
            break

        people.append(
            Person(
                sheet.cell(row=row, column=1).value,
                sheet.cell(row=row, column=2).value,
                sheet.cell(row=row, column=3).value,
                sheet.cell(row=row, column=4).value
            )
        )
        row += 1
    return people

def read_motorbikes(workbook: openpyxl.Workbook, sheet_name: str = "motorbikes", heading: bool = True) -> list[Motorbike]:
    sheet = workbook[sheet_name if sheet_name is not None else "motorbikes"]
    motorbikes = []

    row = 2 if heading else 1
    while True:
        cell = sheet.cell(row=row, column=1)
        if cell.value is None:
            break

        motorbikes.append(Motorbike(
            sheet.cell(row=row, column=1).value,
            sheet.cell(row=row, column=2).value,
            int(sheet.cell(row=row, column=3).value),
            bool(sheet.cell(row=row, column=4).value)
        ))
        row += 1

    return motorbikes


def write_motorbikes(motorbikes: list[Motorbike], workbook: openpyxl.Workbook, sheet_name: str = "motorbikes", heading: bool = True) -> None:
    sheet = workbook.create_sheet(sheet_name if sheet_name is not None else "cars")

    if heading:
        field_names = ["plate", "type", "year", "automatic", "owner id"]
        for col in range(len(field_names)):
            sheet.cell(row=1, column=col + 1, value=field_names[col])

    offset = 2 if heading else 1
    for row in range(len(motorbikes)):
        sheet.cell(row=row + offset, column=1, value=motorbikes[row].plate)
        sheet.cell(row=row + offset, column=2, value=motorbikes[row].type)
        sheet.cell(row=row + offset, column=3, value=motorbikes[row].year)
        sheet.cell(row=row + offset, column=4, value=motorbikes[row].automatic)
        sheet.cell(row=row + offset, column=5, value=motorbikes[row].owner_id)


def read_streets(workbook: openpyxl.Workbook, sheet_name: str = "streets", heading: bool = True) -> list[Street]:
    sheet = workbook[sheet_name if sheet_name is not None else "street"]
    airports = []

    row = 2 if heading else 1
    while True:
        cell = sheet.cell(row=row, column=1)
        if cell.value is None:
            break

        streets.append(Streets(
            sheet.cell(row=row, column=1).value,
            sheet.cell(row=row, column=2).value,
            sheet.cell(row=row, column=3).value,
            sheet.cell(row=row, column=4).value,
            sheet.cell(row=row, column=4).value
        ))
        row += 1

    return streets


def write_streets(streets: list[Street], workbook: openpyxl.Workbook, sheet_name: str = "streets",
                   heading: bool = True) -> None:
    sheet = workbook.create_sheet(sheet_name if sheet_name is not None else "streets")

    if heading:
        field_names = ["code", "name", "city", "state", "country"]
        for col in range(len(field_names)):
            sheet.cell(row=1, column=col + 1, value=field_names[col])

    offset = 2 if heading else 1
    for row in range(len(streets)):
        sheet.cell(row=row + offset, column=1, value=streets[row].code)
        sheet.cell(row=row + offset, column=2, value=streets[row].name)
        sheet.cell(row=row + offset, column=3, value=streets[row].city)
        sheet.cell(row=row + offset, column=4, value=streets[row].state)
        sheet.cell(row=row + offset, column=5, value=streets[row].country)


def read(entity_type: Type[object], workbook: openpyxl.Workbook,
         sheet_name: str = None, heading: bool = True) -> list[object]:
    if entity_type == Person:
        return read_people(workbook, sheet_name=sheet_name, heading=heading)
    elif entity_type == Motorbike:
        return read_motorbikes(workbook, sheet_name=sheet_name, heading=heading)
    elif entity_type == Street:
        return read_streets(workbook, sheet_name=sheet_name, heading=heading)
    else:
        raise RuntimeError("Unknown type of entity")


def write(entities: list[object], workbook: openpyxl.Workbook,
          sheet_name: str = None, heading: bool = True) -> None:
    if isinstance(entities[0], Person):
        return write_people([cast(Person, e) for e in entities], workbook, sheet_name=sheet_name, heading=heading)
    elif isinstance(entities[0], Motorbike):
        return write_motorbikes([cast(Motorbike, e) for e in entities], workbook, sheet_name=sheet_name, heading=heading)
    elif isinstance(entities[0], Street):
        return write_streets([cast(Street, e) for e in entities], workbook, sheet_name=sheet_name, heading=heading)
    else:
        raise RuntimeError("Unknown type of entity")


if __name__ == "__main__":
    wb = Workbook()
    write_people(
        generate_people(10),
        wb
    )
    wb.save("D:/people.xlsx")

    wb = openpyxl.load_workbook("D:/people.xlsx")
    for person in read_people(wb):
        print(person)


def write_streets():
    return None